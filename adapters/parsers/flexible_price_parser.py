"""Универсальный парсер Excel прайсов с конфигурируемыми алиасами колонок.

Идея: один парсер для любых прайс-листов с разным набором колонок и их имён.
Каждой роли (article/description/price/...) соответствует упорядоченный список
ключевых слов (алиасов). Первое совпадение в заголовке листа выигрывает.

Обязательные роли (без них лист пропускается): article, price.
Опциональные (по умолчанию): description, units, manufacturer, code_1c, quantity, currency.
Лишние колонки игнорируются.

Использование:
    parser = FlexiblePriceParser({
        'column_aliases': {...},        # опционально, есть DEFAULT_ALIASES
        'required_columns': {'article', 'price'},
        'optional_columns': {'description', 'units', ...},
        'vendor_from_column': True,     # брать вендор из колонки manufacturer
        'default_vendor': 'GENERIC',    # фолбэк, если колонки manufacturer нет
        'sheet_name': 'Лист1',          # None → перебрать все листы
        'header_row': 0,                # None → автопоиск
        'max_header_scan': 10,
    }, normalizer)
    items = parser.parse(Path('price.xlsx'), vendor='CHINT')
"""
import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl

from domain.entities.price_item import PriceItem
from domain.interfaces.parser import IParser
from domain.services.data_normalizer import DataNormalizer
from utils.normalizer import ArticleNormalizer

logger = logging.getLogger(__name__)


# Стандартные алиасы колонок (приоритет: первое совпадение выигрывает).
# Все сравнения — case-insensitive substring match.
# Важно: «код» в прайсе обычно = код 1С (ArticlePC), а не артикул.
# А «Артикул РС» — тоже код 1С. Поэтому «код» и «артикул рс» живут в code_1c,
# а не в article. Распределение ролей решается приоритетом ролей в _map_columns.
DEFAULT_ALIASES: Dict[str, List[str]] = {
    'article': ['артикул'],
    'description': [
        'наименование для печати',  # приоритет у "для печати"
        'наименование',
        'полное описание',
        'описание',
    ],
    'price': [
        'тариф с ндс',
        'тариф',
        'цена с ндс',
        'цена',
        'без ндс',
    ],
    'units': [
        'единица хранения',
        'ед. изм',
        'ед изм',
        'ед.',
        'единица измерения',
    ],
    'manufacturer': [
        'производитель',
        'бренд',
        'вендор',
        'марка',
    ],
    'code_1c': [
        'артикул рс',  # самый специфичный — должен перебить «артикул» в article
        'код 1с',
        'код1с',
        'артикул 1с',
        'код',         # голое «Код» в прайсе = код 1С (ArticlePC)
    ],
    'quantity': [
        'кол-во',
        'количество',
    ],
    'currency': [
        'валюта',
    ],
}


# Роли, без которых лист считается "без заголовков" и пропускается.
DEFAULT_REQUIRED: Set[str] = {'article', 'price'}

# Роли, которые могут отсутствовать без падения парсинга.
DEFAULT_OPTIONAL: Set[str] = {
    'description', 'units', 'manufacturer', 'code_1c', 'quantity', 'currency',
}

# Порядок приоритета ролей при маппинге колонок: более специфичные проверяются
# раньше и помечают колонку "занятой". Это нужно, чтобы «Артикул РС» ушёл в
# code_1c, а не в article (т.к. содержит подстроку «артикул»).
ROLE_PRIORITY: List[str] = [
    'code_1c',       # «Артикул РС» / «Код 1С» / «Код» — сначала
    'article',       # затем «Артикул»
    'description',
    'price',
    'units',
    'manufacturer',
    'quantity',
    'currency',
]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip().lower()


def _find_column(
    headers: List[str],
    keywords: List[str],
    used: Optional[Set[int]] = None,
) -> Optional[int]:
    """Возвращает индекс первой колонки, заголовок которой содержит
    хотя бы одно ключевое слово (case-insensitive substring).

    Параметр `used` — множество уже занятых индексов; такие колонки
    пропускаются. Это нужно для приоритизации ролей: если «Код» уже
    ушёл в code_1c, то article его не перебьёт.
    """
    skipped = used or set()
    for idx, header in enumerate(headers):
        if idx in skipped:
            continue
        h = _normalize_header(header)
        if not h:
            continue
        if any(kw in h for kw in keywords):
            return idx
    return None


def _is_price_on_request(raw: Any) -> bool:
    if raw is None:
        return False
    return 'запрос' in str(raw).strip().lower()


def _clean_price(raw: Any) -> Optional[Decimal]:
    if raw is None:
        return None
    try:
        value = str(raw).strip().replace(' ', '').replace(',', '.')
        d = Decimal(value)
        return d if d >= 0 else None
    except (InvalidOperation, ValueError):
        return None


class FlexiblePriceParser(IParser):
    """Generic Excel-парсер прайсов.

    Ищет колонки по спискам алиасов (приоритет: первое совпадение).
    Поддерживает несколько листов: парсит все, где найдены обязательные роли.
    Дедуплицирует по (vendor, article) — побеждает первый встреченный.
    """

    def __init__(
        self,
        config: Optional[Dict[str, Any]] = None,
        normalizer: Optional[ArticleNormalizer] = None,
    ):
        cfg = config or {}
        self.aliases: Dict[str, List[str]] = cfg.get('column_aliases', DEFAULT_ALIASES)
        self.required: Set[str] = set(cfg.get('required_columns', DEFAULT_REQUIRED))
        self.optional: Set[str] = set(cfg.get('optional_columns', DEFAULT_OPTIONAL))
        self.vendor_from_column: bool = cfg.get('vendor_from_column', False)
        self.default_vendor: str = cfg.get('default_vendor', 'GENERIC')
        self.sheet_name: Optional[str] = cfg.get('sheet_name')
        self.header_row: Optional[int] = cfg.get('header_row')
        self.max_header_scan: int = cfg.get('max_header_scan', 10)
        self.skip_empty_description: bool = cfg.get('skip_empty_description', True)

        self.normalizer = normalizer or ArticleNormalizer()
        self.data_normalizer = DataNormalizer()
        self._last_stats: Dict[str, int] = {'total': 0, 'with_headers': 0}

    def parse(self, file_path: Path, vendor: str) -> List[PriceItem]:
        logger.info(
            "[FLEX] FlexiblePriceParser.parse: vendor=%s file=%s",
            vendor, file_path,
        )
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        items: List[PriceItem] = []
        seen: Set[Tuple[str, str]] = set()

        sheets = [self.sheet_name] if self.sheet_name else list(wb.sheetnames)
        self._last_stats = {'total': len(sheets), 'with_headers': 0}

        for name in sheets:
            if name not in wb.sheetnames:
                logger.debug("[FLEX] лист '%s' не найден в книге", name)
                continue
            sheet_items, has_headers = self._parse_sheet(wb[name], vendor, seen)
            if has_headers:
                self._last_stats['with_headers'] += 1
            items.extend(sheet_items)

        wb.close()
        logger.info(
            "[FLEX] итого %d позиций, листов: %d всего, %d с заголовками",
            len(items), self._last_stats['total'], self._last_stats['with_headers'],
        )
        return items

    def _find_header_row(self, rows: List[Tuple]) -> Optional[int]:
        """Ищет строку заголовков среди первых max_header_scan строк.

        Критерий: на строке должны совпасть ключевые слова минимум для двух
        обязательных ролей (по умолчанию article + price). Остальные роли
        ищутся на той же строке.
        """
        limit = min(self.max_header_scan, len(rows))
        for idx in range(limit):
            cells = [_normalize_header(c) for c in rows[idx]]
            hits = 0
            for role in self.required:
                kws = self.aliases.get(role, [])
                if any(any(kw in c for kw in kws) for c in cells):
                    hits += 1
            if hits >= 2:
                return idx
        return None

    def _parse_sheet(
        self,
        ws,
        default_vendor: str,
        seen: Set[Tuple[str, str]],
    ) -> Tuple[List[PriceItem], bool]:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], False

        header_idx = (
            self.header_row
            if self.header_row is not None
            else self._find_header_row(rows)
        )
        if header_idx is None:
            logger.debug("[FLEX] лист '%s' — строка заголовков не найдена", ws.title)
            return [], False

        header_row = [_normalize_header(c) for c in rows[header_idx]]
        col_map = self._map_columns(header_row)

        missing = [r for r in self.required if col_map.get(r) is None]
        if missing:
            logger.debug(
                "[FLEX] лист '%s' — не найдены обязательные колонки %s, пропускаем",
                ws.title, missing,
            )
            return [], False

        logger.info(
            "[FLEX] лист '%s' заголовки на строке %d, col_map=%s",
            ws.title, header_idx + 1,
            {k: v for k, v in col_map.items() if v is not None},
        )

        items: List[PriceItem] = []
        for row in rows[header_idx + 1:]:
            item = self._parse_row(row, col_map, default_vendor, seen)
            if item is not None:
                items.append(item)

        return items, True

    def _map_columns(self, headers: List[str]) -> Dict[str, Optional[int]]:
        """Маппит заголовки на роли с учётом приоритета ролей.

        Сначала проверяются более специфичные роли (code_1c раньше article),
        и каждая найденная колонка помечается занятой — более поздние роли
        её уже не подхватят. Это решает конфликт «Артикул РС» vs «Артикул»:
        первый уходит в code_1c, второй — в article.
        """
        result: Dict[str, Optional[int]] = {}
        used: Set[int] = set()
        for role in ROLE_PRIORITY:
            kws = self.aliases.get(role, [])
            if not kws:
                continue
            idx = _find_column(headers, kws, used=used)
            result[role] = idx
            if idx is not None:
                used.add(idx)
        return result

    def map_columns(self, headers: List[str]) -> Dict[str, Optional[int]]:
        """Публичная обёртка над _map_columns — для отладки и тестов.

        Принимает список заголовков колонок (case-insensitive),
        возвращает словарь {роль: индекс колонки или None}.
        """
        return self._map_columns([_normalize_header(h) for h in headers])

    def _parse_row(
        self,
        row: Tuple,
        col_map: Dict[str, Optional[int]],
        default_vendor: str,
        seen: Set[Tuple[str, str]],
    ) -> Optional[PriceItem]:
        def cell(role: str) -> Optional[str]:
            idx = col_map.get(role)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            s = str(v).strip()
            return s if s else None

        article_raw = cell('article')
        if not article_raw:
            return None

        article = self.data_normalizer.normalize_article(article_raw, default_vendor)
        if not article or article in ('NAN', 'NONE'):
            return None

        price_idx = col_map.get('price')
        raw_price = row[price_idx] if price_idx is not None and price_idx < len(row) else None
        price = _clean_price(raw_price)
        if price is None:
            if _is_price_on_request(raw_price):
                price = Decimal(0)
            else:
                return None

        description = cell('description') or ''
        if self.skip_empty_description and not description:
            return None

        # Вендор: из колонки manufacturer (если разрешено и есть), иначе переданный.
        if self.vendor_from_column:
            vendor = (cell('manufacturer') or default_vendor).strip() or default_vendor
        else:
            vendor = default_vendor

        units_raw = cell('units')
        units = self.data_normalizer.normalize_unit(units_raw) if units_raw else 'шт'

        key = (vendor.upper(), article)
        if key in seen:
            return None
        seen.add(key)

        description = self.data_normalizer.normalize_description(description)

        # Код 1С (опционально). Достаётся, только если колонка найдена
        # в шапке; иначе остаётся пустым и в БД не пишется.
        code_1c = (cell('code_1c') or '').strip()

        try:
            return PriceItem(
                vendor=vendor,
                article=article,
                description=description,
                price=price,
                units=units,
                storage='',
                code_1c=code_1c,
            )
        except ValueError as e:
            logger.debug(
                "[FLEX] пропускаем строку article=%s: %s", article, e,
            )
            return None

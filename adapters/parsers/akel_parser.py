import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional, Dict

import openpyxl

from domain.entities.price_item import PriceItem
from domain.interfaces.parser import IParser

logger = logging.getLogger(__name__)

# Нечёткие ключевые слова для поиска нужных колонок
_COLUMN_KEYWORDS = {
    'article': ['артикул'],
    'description': ['полное описание', 'описание'],
    'storage': ['складской статус', 'статус'],
    'price': ['тариф', 'без ндс'],
}


def _find_column(headers: List[str], keywords: List[str]) -> Optional[int]:
    """Возвращает индекс первой колонки, название которой содержит все ключевые слова."""
    for idx, header in enumerate(headers):
        if header is None:
            continue
        h = str(header).strip().lower()
        if all(kw in h for kw in keywords):
            return idx
    return None


def _is_price_on_request(raw) -> bool:
    """Проверяет, является ли сырое значение цены 'по запросу'."""
    if raw is None:
        return False
    return 'запрос' in str(raw).strip().lower()


def _clean_price(raw) -> Optional[Decimal]:
    """Конвертирует сырое значение цены в Decimal. Возвращает None если не получилось."""
    if raw is None:
        return None
    try:
        value = str(raw).strip().replace(' ', '').replace(',', '.')
        d = Decimal(value)
        return d if d >= 0 else None
    except (InvalidOperation, ValueError):
        return None


class AkelParser(IParser):
    """Парсер многолистового Excel прайса AKEL.

    Логика:
    - Заголовки на строке 2 (индекс 1 в 0-based openpyxl = row 2).
    - Данные начиная со строки 3 (row 3+).
    - Колонки определяются нечётким поиском.
    - Листы без колонок article + price пропускаются.
    - Дубликаты артикулов дедуплицируются (берётся первый встреченный).
    """

    def parse(self, file_path: Path, vendor: str) -> List[PriceItem]:
        logger.info("[FIX] AkelParser.parse: vendor=%s file=%s", vendor, file_path)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        items: List[PriceItem] = []
        seen_articles: set = set()
        total_sheets = len(wb.sheetnames)
        sheets_with_headers = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_items, has_headers = self._parse_sheet(ws, sheet_name, vendor, seen_articles)
            if has_headers:
                sheets_with_headers += 1
            items.extend(sheet_items)

        wb.close()
        self._last_stats = {'total': total_sheets, 'with_headers': sheets_with_headers}
        logger.info(
            "[FIX] AkelParser: итого распознано %d позиций, листов: %d всего, %d с заголовками",
            len(items), total_sheets, sheets_with_headers,
        )
        return items

    def _parse_sheet(self, ws, sheet_name: str, vendor: str, seen_articles: set):
        """Возвращает (items, has_headers) где has_headers — найдены ли обязательные колонки."""
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 3:
            logger.debug("[FIX] AkelParser: лист '%s' содержит < 3 строк — пропускаем", sheet_name)
            return [], False

        # Заголовки на строке 2 (индекс 1)
        header_row = [str(h).strip().lower() if h is not None else '' for h in rows[1]]

        col_map = self._map_columns(header_row)

        if col_map.get('article') is None or col_map.get('price') is None:
            logger.debug(
                "[FIX] AkelParser: лист '%s' — не найдены обязательные колонки (article=%s, price=%s) — пропускаем",
                sheet_name, col_map.get('article'), col_map.get('price')
            )
            return [], False

        logger.debug("[FIX] AkelParser: лист '%s' col_map=%s", sheet_name, col_map)

        items: List[PriceItem] = []
        for row in rows[2:]:  # данные с row 3
            item = self._parse_row(row, col_map, vendor, sheet_name, seen_articles)
            if item is not None:
                items.append(item)

        logger.debug("[FIX] AkelParser: лист '%s' → %d позиций", sheet_name, len(items))
        return items, True

    def _map_columns(self, headers: List[str]) -> Dict[str, Optional[int]]:
        return {
            'article': _find_column(headers, _COLUMN_KEYWORDS['article']),
            'description': _find_column(headers, _COLUMN_KEYWORDS['description'])
                           or _find_column(headers, ['описание']),
            'storage': _find_column(headers, _COLUMN_KEYWORDS['storage'])
                       or _find_column(headers, ['статус']),
            'price': _find_column(headers, _COLUMN_KEYWORDS['price']),
        }

    def _parse_row(self, row, col_map: Dict[str, Optional[int]], vendor: str, sheet_name: str, seen_articles: set) -> Optional[PriceItem]:
        def get(col_key) -> Optional[str]:
            idx = col_map.get(col_key)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        article = get('article')
        if not article:
            return None

        if article in seen_articles:
            return None
        seen_articles.add(article)

        raw_price_idx = col_map.get('price')
        raw_price_val = row[raw_price_idx] if raw_price_idx is not None and raw_price_idx < len(row) else None
        price_val = _clean_price(raw_price_val)
        if price_val is None:
            if _is_price_on_request(raw_price_val):
                price_val = Decimal(0)
            else:
                return None

        description = get('description') or ''
        storage = get('storage') or ''

        try:
            return PriceItem(
                vendor=vendor,
                article=article,
                description=description,
                price=price_val,
                units='шт',
                storage=storage,
            )
        except ValueError as e:
            logger.debug("[FIX] AkelParser: пропускаем строку sheet='%s' article=%s: %s", sheet_name, article, e)
            return None

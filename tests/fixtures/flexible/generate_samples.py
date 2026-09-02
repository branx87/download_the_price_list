"""Генератор тестовых xlsx-фикстур для FlexiblePriceParser.

Запускать из корня проекта:
    python tests/fixtures/flexible/generate_samples.py

Создаёт три файла:
  - sample_format1.xlsx — колонки: Наименование (для печати) | Артикул | Код | Ед. хранения | Производитель
  - sample_format2.xlsx — колонки: Производитель | Артикул | Наименование | Кол-во | Ед. изм | Тариф | Валюта
  - sample_mixed.xlsx  — оба листа в одной книге, плюс лишние колонки
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# Принудительно UTF-8 для вывода в консоль Windows (cp1251 по умолчанию).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

FIXTURES_DIR = Path(__file__).parent


def _header_row(ws, headers, freeze=True):
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = Font(bold=True)
    if freeze:
        ws.freeze_panes = 'A2'


def build_format1(path: Path) -> None:
    """Колонки: Наименование (приоритет «для печати») | Артикул | Код | Ед. хранения | Производитель."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Прайс'
    _header_row(ws, [
        'Наименование для печати',
        'Артикул',
        'Код',
        'Единица хранения',
        'Производитель',
        'Комментарий',  # лишняя колонка — должна игнорироваться
    ])
    rows = [
        ('Контактор КМИ-22510 25А 230В/AC3 1НО', 'KMI-22510', '00-000001', 'шт', 'IEK', 'склад МСК'),
        ('Автомат ВА47-29 1Р 16А 4.5кА',          'MVA20-1-016', '00-000002', 'шт', 'IEK', ''),
        ('Розетка РА16-046Б',                      'RA16-046B',   '00-000003', 'шт', 'IEK', ''),
        ('DIN-рейка 35см оцинкованная',            'DIN-35-300',  '00-000004', 'м',  'IEK', 'по запросу'),
    ]
    for r in rows:
        # последняя цена отсутствует — это «по запросу» (колонка цены вообще не выведена)
        ws.append([*r, 'по запросу' if r[0].startswith('DIN') else 0])
    # добавим колонку цены в конец для большинства строк
    # (для наглядности: 4-я колонка цены, но сначала идёт «Комментарий» — его игнорируем)
    # Перепишем проще: добавим ещё одну колонку
    ws.cell(row=1, column=7, value='Цена с НДС, руб.').font = Font(bold=True)
    prices = [1250.50, 380.00, 540.00, 0.00]  # DIN — по запросу
    for i, price in enumerate(prices, start=2):
        ws.cell(row=i, column=7, value=price)
    wb.save(path)


def build_format2(path: Path) -> None:
    """Колонки: Производитель | Артикул | Наименование | Кол-во | Ед. изм | Тариф | Валюта."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    _header_row(ws, [
        'Производитель',
        'Артикул',
        'Наименование',
        'Кол-во',
        'Ед. изм',
        'Тариф',
        'Валюта',
        'Срок',  # лишняя колонка
    ])
    rows = [
        ('TDM',  'TDM-SQ1202-0024',  'Автомат 1Р 16А',     1, 'шт', 410.00, 'RUB', '5 дн.'),
        ('TDM',  'TDM-SQ1202-0032',  'Автомат 1Р 25А',     1, 'шт', 420.00, 'RUB', '5 дн.'),
        ('TDM',  'TDM-НН-100',       'Наконечник НН 100', 100, 'упак', 8500.00, 'RUB', '2 дн.'),
        # 'запрос' ставим именно в колонку «Тариф» (цена) — проверяем, что парсер
        # распознаёт «по запросу» и обнуляет цену. В «Срок» — обычное значение.
        ('TDM',  'TDM-МК-10',        'Маркер кабельный',  50, 'упак', 'запрос', 'RUB', '5 дн.'),
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)


def build_mixed(path: Path) -> None:
    """Две страницы (формат 1 + формат 2) и лист-«пустышка» без обязательных колонок."""
    wb = openpyxl.Workbook()

    # Лист 1 — формат 1
    ws1 = wb.active
    ws1.title = 'Формат1'
    _header_row(ws1, ['Наименование', 'Артикул', 'Производитель', 'Цена', 'Ед.'])
    for r in [
        ('Контактор КМИ 25А', 'KMI-22510', 'IEK', 1250.0, 'шт'),
        ('Контактор КМИ 40А', 'KMI-22540', 'IEK', 2400.0, 'шт'),
    ]:
        ws1.append(r)

    # Лист 2 — формат 2
    ws2 = wb.create_sheet('Формат2')
    _header_row(ws2, ['Производитель', 'Артикул', 'Наименование', 'Ед. изм', 'Тариф', 'Валюта'])
    for r in [
        ('TDM', 'TDM-001', 'Изолента', 'шт', 95.0, 'RUB'),
        ('TDM', 'TDM-002', 'Стяжка 200мм', 'упак', 350.0, 'RUB'),
    ]:
        ws2.append(r)

    # Лист 3 — «пустышка» без обязательных колонок
    ws3 = wb.create_sheet('Служебный')
    _header_row(ws3, ['ID', 'Комментарий'])
    ws3.append([1, 'этот лист должен быть пропущен'])

    wb.save(path)


def main() -> int:
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    targets = {
        'sample_format1.xlsx': build_format1,
        'sample_format2.xlsx': build_format2,
        'sample_mixed.xlsx':  build_mixed,
    }
    for name, builder in targets.items():
        path = FIXTURES_DIR / name
        builder(path)
        print(f'[OK] {name}')
    print(f'\nГотово. Файлы в {FIXTURES_DIR}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
